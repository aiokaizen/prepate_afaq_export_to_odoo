import os
import logging
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer import excel

from django.utils.text import slugify

from src.settings import BASE_DIR


def generate_number_with_fixed_size(num, size=5):
    count = len(str(num))
    return ''.join(['0' for i in range(size - count)]) + str(num)


def get_category(category, categories):
    res = list(filter((lambda c : c[1] == category), categories))
    if len(res) > 0:
        return res[0]
    return None


def get_cell_str(col, row):
    return f"{get_column_letter(col)}{row}"


def handle_excel_file(
    file_name,
    start_row=1,
    extract_columns=None,
    target_sheet=None,
):
    """
    file_name | string: The path to open or a File like object
    start_row | int: the number of row where the header of the file is located.
    extract_columns | List of strings: the names of columns to extract from the file.
    The extract_columns param will be slugified as well as the columns from the excel file,
    so caps, spaces, and special characters are ignored, making it easier to match.
    target_sheet | string: Name of the target sheet

    example:
    >>> start_row = 1
    >>> column_names = [
    >>>     "name", "age", "address"
    >>> ]
    >>> data = handle_excel_file("file.xlsx", start_row, column_names)
    """

    if type(start_row) != int or start_row <= 0:
        raise Exception("'start_row' attribute is invalid!")
    start_row -= 1

    try:
        # Slugify extract_columns
        if not extract_columns:
            extract_columns = []
            slugified_extract_columns = []
        else:
            slugified_extract_columns = [
                slugify(name) for name in extract_columns
            ]

        wb = openpyxl.load_workbook(file_name)
        if target_sheet:
            worksheet = wb[target_sheet]
        else:
            worksheet = wb.active

        # Extract columns names from the excel file
        all_column_names = []
        columns_indexes = []

        fill_extract_columns = False if slugified_extract_columns else True

        for index, row in enumerate(worksheet.iter_rows()):
            if index != start_row:
                continue

            for col_index, cell in enumerate(row):
                value = cell.value
                if not value:
                    break

                slugified_value = slugify(str(value))
                if fill_extract_columns:
                    extract_columns.append(str(value))
                    slugified_extract_columns.append(slugified_value)
                all_column_names.append(slugified_value)
                columns_indexes.append(col_index)

        # Check if all extract_columns exist in the excel file.
        for column_name in slugified_extract_columns:
            if column_name not in all_column_names:
                raise Exception(
                    "The uploaded file does not contain a column named '%s'."
                    % (column_name,)
                )

        extract_columns_indexes = []
        for index, column_name in enumerate(all_column_names):
            if column_name in slugified_extract_columns:
                extract_columns_indexes.append(index)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for index, row in enumerate(worksheet.iter_rows()):
            if index <= start_row:
                continue

            is_row_empty = True
            row_data = dict()
            for col_index, cell in enumerate(row):
                if col_index not in extract_columns_indexes:
                    continue

                row_data[all_column_names[col_index]] = cell.value
                if cell.value:
                    is_row_empty = False

            if not is_row_empty:
                excel_data.append(row_data)

        # return odict object with the following format:
        return dict(
            column_names=slugified_extract_columns,
            column_names_display=extract_columns,
            data=excel_data,
        )
    except Exception as e:
        logging.error(e)
        return None


def export_xlsx(
    data: dict,
    sheet_title="Data",
    freeze_header=True,
):
    """
    This function creates a tmp .xlsx file based on the data provided and
    returns a result that has the tmp file path in it's instance attribute.

    :data:dict:Data to be exported to excel file.
    :sheet_title:str:The title of the generated spread sheet.
    :freeze_header:bool:If set to True (default), the header aread will be fixed on scroll.
    """

    start_row_index = 1
    start_col_index = 1
    saved = None

    col_titles = data["col_titles"]

    try:
        # Workbook initialization
        wb = Workbook()

        # Sheet setup
        ws = wb.active
        ws.title = sheet_title

        # Freeze the table header
        if freeze_header:
            ws.freeze_panes = ws.cell(row=start_row_index + 1, column=1)

        for index, col_title in enumerate(col_titles):
            ws.cell(
                row=start_row_index, column=index + start_col_index, value=col_title
            )

        for row_index, row in enumerate(data["data"]):
            for col_index, value in enumerate(row):
                ws.cell(
                    row=row_index + start_row_index + 1,
                    column=col_index + start_col_index,
                    value=value,
                )

        export_timestamp = int(datetime.now().strftime("%Y%m%d%H%M%S"))
        filename = f"{slugify(sheet_title)}_{export_timestamp}.xlsx"
        tmp_file_path = os.path.join(BASE_DIR, "export", filename)

        saved = excel.save_workbook(wb, tmp_file_path)
    except Exception as e:
        logging.error(e)
        return None

    if not saved:
        print("The tmp file could not be created!")

    return print("File exported successfully", tmp_file_path)
